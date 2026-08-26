"""
네이버 지도에서 지역+음식종류로 검색한 뒤, 검색 결과 목록을 스크롤해서 모으고
브랜드명/카테고리/리뷰수/대표 리뷰/지도링크를 전부 그 목록 화면에서만 뽑아
엑셀로 저장하는 스크립트. 개별 가게의 상세 페이지에는 들어가지 않는다
(하나씩 들어가면 훨씬 느려지고, 목록에 이미 필요한 정보가 다 있다).

기준 리뷰 수를 넘는 가게는 목록에서 클릭해 옆에 뜨는 상세 패널까지는 확인한다.
목록 카드 자체에 리뷰 문장이 안 보이는 가게(미쉐린 등 일부 고급 음식점에서
자주 보임)는 상세 패널을 스크롤해서 'AI 브리핑' 요약으로 대신 채운다
(_parse_ai_briefing, _read_ai_briefing_snippet 참고).

_try_get_extra_info, get_extra_info 등 별도 검색으로 상세 페이지에 다시
들어가는 코드는 남겨두었지만 기본 흐름에서는 쓰지 않는다. 필요하면 나중에
다시 불러 쓸 수 있다.

네이버 지역검색 API는 더 이상 쓰지 않는다 (검색어당 5개 제한이 있고, 어차피
목록 화면에 리뷰 수가 그대로 보이기 때문). API 키도 필요 없다.

실행 방법은 README.md를 참고하세요.
모든 설정값(지역, 찾을 음식 종류, 최소 리뷰 수 등)은 config.py에서 바꿉니다.
"""

import os
import re
import time
import random
from datetime import datetime
from urllib.parse import quote

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config


# 브라우저 로그인/쿠키 정보를 저장해두는 폴더.
# 매번 새 브라우저(신규 방문자)인 것처럼 접속하면 네이버가 자동화로 의심하기 쉬워서,
# 같은 브라우저 프로필을 계속 재사용해 "이전에도 왔던 사람"처럼 보이게 한다.
PROFILE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "browser_profile")

# 최종 엑셀 컬럼 순서
HEADERS = ["브랜드명", "키워드", "카테고리", "리뷰수", "대표 리뷰", "네이버지도주소"]

# 엑셀 서식에 쓸 값들
FONT_NAME = "나눔바른고딕OTF"
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


class _NullStopEvent:
    """중지 신호를 안 받는 경우(터미널에서 바로 실행할 때 등) 쓰는 빈 자리표시자.

    threading.Event()와 똑같이 .is_set()을 갖고 있어서, 중지 기능이 있는
    launcher.py에서 넘겨준 진짜 이벤트든, 이 빈 이벤트든 코드에서 똑같이 다룰 수 있다.
    """

    def is_set(self):
        return False


_NO_STOP = _NullStopEvent()

# 카테고리 추출용 디버그를 몇 번 보여줬는지 (임시 디버그용, 실행할 때마다 초기화되지는 않음)
_category_debug_shown = [0]

# 목록 카드에 '리뷰' 글자 자체가 없어서 건너뛴 항목의 디버그를 몇 번 보여줬는지.
# (브랜드명으로 검색하면 목록 카드에 리뷰 수가 안 보이는 경우가 있을 수 있어서,
# 이런 경우를 진단하기 위한 디버그. 실제로 이런 사례를 만나면 이 출력을 보고
# 카드 글자 형태에 맞춰 다시 고쳐야 한다.)
_no_review_debug_shown = [0]


def _district_only(address):
    """'서울 강남구 도산대로1길 10' -> '서울 강남구'처럼 구까지만 남긴다."""
    parts = address.split()
    return " ".join(parts[:2]) if len(parts) >= 2 else address


def _parse_review_count(text):
    """'리뷰 1.1만', '리뷰 1,234' 같은 표기에서 숫자를 뽑아낸다.

    네이버 플레이스는 리뷰가 많으면 '1.1만'처럼 만/천 단위로 줄여서 보여준다.
    """
    match = re.search(r"리뷰\s*([\d,]+(?:\.\d+)?)\s*(만|천)?", text)
    if not match:
        return None

    value = float(match.group(1).replace(",", ""))
    unit = match.group(2)
    if unit == "만":
        value *= 10000
    elif unit == "천":
        value *= 1000

    return int(value)


# 이름/카테고리에 공백 없이 바로 붙어 나올 수 있는 '예약/톡톡/쿠폰/새로오픈' 같은
# 배지 글자들. 목록 항목이나 상세 패널 글자에서 이런 배지가 섞여 나오면 잘라낸다.
_KNOWN_BADGE_WORDS = ["예약", "톡톡", "쿠폰", "포장주문", "발견", "광고", "새로오픈"]


def _strip_trailing_badges(name):
    changed = True
    while changed:
        changed = False
        for badge in _KNOWN_BADGE_WORDS:
            if name.endswith(badge):
                name = name[: -len(badge)]
                changed = True
    return name.strip()


def _strip_leading_badges(text):
    changed = True
    while changed:
        changed = False
        for badge in _KNOWN_BADGE_WORDS:
            if text.startswith(badge):
                text = text[len(badge):]
                changed = True
    return text.strip()


# 목록 항목 텍스트에서 리뷰 글(자유 문장)이 아닌, 구조적으로 나오는 문구들.
# 리뷰 글을 뽑을 때 이런 줄은 제외한다.
_STRUCTURAL_HINTS = (
    "혜택", "쿠폰", "포인트", "리뷰", "예약", "톡톡", "광고", "connect+", "포장주문",
    # 미쉐린 등 일부 카드는 리뷰 문장 대신 영업시간 상태만 나온다.
    "영업 중", "영업 전", "영업 시작", "브레이크타임", "라스트오더",
)

# 리뷰 문장이 아니라 카드에 고정으로 붙는 버튼/배지 문구들 (줄 전체가 이거랑
# 똑같으면 제외한다).
_STRUCTURAL_EXACT_LINES = {
    "상세주소 열기", "출발도착", "현재 위치에서", "플레이스 플러스", "네이버페이",
    "휠체어 출입 가능",
}

# '서울 마포구 합정동'처럼 지역명만 딱 나오는 주소 줄 (리뷰가 아니라 주소다).
_ADDRESS_LINE_RE = re.compile(
    r"^(서울|부산|대구|인천|광주|대전|울산|세종|경기|강원|충북|충남|전북|전남|경북|경남|제주)"
    r"\S*\s+\S+(시|군|구)\s+\S+(동|읍|면|가|리)$"
)

# '6.7km'처럼 현재 위치에서의 거리만 나오는 줄.
_DISTANCE_LINE_RE = re.compile(r"^\d+(\.\d+)?km$")

# 대표 리뷰 추출용 디버그를 몇 번 보여줬는지 (임시 디버그용)
_review_debug_shown = [0]

# AI 브리핑으로 대신 채운 경우를 몇 번 보여줬는지 (임시 디버그용)
_ai_briefing_debug_shown = [0]


def _extract_review_snippet(lines):
    """목록 카드 안에 보이는 짧은 리뷰 후기 한 줄을 뽑아낸다.

    카드 아래쪽에 실제 리뷰어가 쓴 문장이 그대로 나오는 경우가 많다. 이름/카테고리/
    배지/리뷰수/영업시간/주소/거리/편의시설 태그 같은 '구조적인' 줄들을 제외하고
    남는 자유 문장 중, 말줄임표(...)로 끝나지 않는(=안 잘린) 문장을 우선으로
    마지막(카드 아래쪽에 가까운) 것을 고른다.

    카드에 따라(예: 미쉐린 가이드 등) 리뷰 문장 자체가 아예 없이 영업시간/주소/
    거리/편의시설 정보만 나오는 경우가 있다 - 이런 카드는 후보가 하나도 안 남아서
    빈 문자열을 돌려주는 게 맞다 (호출하는 쪽에서 AI 브리핑으로 대신 채운다).
    """
    candidates = []
    for l in lines[1:]:  # lines[0]은 가게 이름이라 리뷰 후보에서 제외한다.
        if len(l) < 8:
            continue
        if any(h in l for h in _STRUCTURAL_HINTS):
            continue
        if l in _STRUCTURAL_EXACT_LINES:
            continue
        if _ADDRESS_LINE_RE.match(l) or _DISTANCE_LINE_RE.match(l):
            continue
        candidates.append(l)

    result = ""
    if candidates:
        complete = [c for c in candidates if not c.endswith("...") and not c.endswith("…")]
        pool = complete if complete else candidates
        result = pool[-1]

    if _review_debug_shown[0] < 5:
        print(f"  ---- (대표 리뷰 디버그) 카드 줄들: {lines!r}")
        print(f"  후보: {candidates!r} -> 선택: {result!r}")
        _review_debug_shown[0] += 1

    return result


def _parse_list_item(text):
    """검색 결과 '목록'에 나오는 항목 하나의 텍스트에서 정보를 뽑아낸다.

    목록의 각 항목은 보통 첫 줄이 가게 이름이고, 어딘가에 '리뷰 N' 형태로
    리뷰 수가 나온다. 리뷰 수가 없는 항목(광고/필터 칩 등)은 걸러진다.

    카테고리/지도 링크는 목록 카드 글자만으로는 정확히 구분하기 어려워서
    (배지/홍보 문구와 뒤섞여 나옴), 여기서는 채우지 않는다 - 기준 리뷰 수를
    넘는 항목만 클릭해서 상세 패널에서 정확하게 가져온다 (_current_list_rows 참고).

    반환값: {"브랜드명", "리뷰수", "대표 리뷰"} 또는 None
    """
    review_count = _parse_review_count(text)
    if review_count is None:
        return None

    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    if not lines:
        return None

    name = _strip_trailing_badges(lines[0])
    if not name:
        return None

    return {
        "브랜드명": name,
        "리뷰수": review_count,
        "대표 리뷰": _extract_review_snippet(lines),
    }


def _extract_name_and_category(body_text):
    """상세 페이지 글자에서 '별점' 바로 앞 두 줄(브랜드명, 카테고리)을 뽑아낸다.

    네이버 플레이스 상세 화면은 보통 '...\\n브랜드명\\n카테고리\\n별점\\n4.85리뷰...'
    순서로 나온다.
    """
    idx = body_text.find("별점")
    if idx == -1:
        return "", ""

    lines = [l.strip() for l in body_text[:idx].split("\n") if l.strip()]
    if not lines:
        return "", ""

    category = lines[-1]
    name = lines[-2] if len(lines) >= 2 else ""
    return name, category


def _extract_address(body_text):
    """상세 페이지 글자에서 '주소' 라벨 바로 다음 줄을 주소로 가져온다."""
    match = re.search(r"주소\n([^\n]+)", body_text)
    return match.group(1).strip() if match else ""


def _parse_ai_briefing(text, max_items=2):
    """'AI 브리핑' 라벨 아래에 나오는 요약 문장을 최대 max_items개까지 가져온다.

    각 요약 문장 끝에는 보통 '닉네임 +2' 같은 출처 표시가 붙는데,
    그 표시를 기준으로 문장 단위를 나눈다.
    """
    match = re.search(r"AI\s*브리핑", text)
    if not match:
        return None, "'AI 브리핑' 글자를 페이지에서 못 찾았습니다 (스크롤해도 안 나타났을 수 있음)."

    section = text[match.end():match.end() + 1500]
    debug_context = f"'AI 브리핑' 뒤 1500자: {section!r}"

    # '다양한 리뷰를 종합해 주요 특징을 요약해 드립니다.' / '주요 특징을 리뷰를
    # 활용해 요약해 드리겠습니다.' 처럼, 안내 문구가 조금씩 다르게 나온다. 정확한
    # 문구를 다 맞추는 대신, 맨 앞 문장이 '요약'이나 '정리'라는 말을 담고 있고
    # '~습니다.'로 끝나면 안내 문구로 보고 통째로 제거한다.
    first_sentence = re.match(r"^(.*?니다\.)\s*", section, flags=re.S)
    if first_sentence and ("요약" in first_sentence.group(1) or "정리" in first_sentence.group(1)):
        section = section[first_sentence.end():]

    # 각 요약 문장은 보통 끝에 '닉네임 +숫자' 형태의 출처 표시가 붙는다.
    bullets = re.findall(r"(.+?)\s*\S+\s*\+\d+", section, flags=re.S)
    bullets = [b.replace("\n", " ").strip() for b in bullets if b.strip()]

    if not bullets:
        # 출처 표시를 못 찾으면, 줄바꿈 기준으로 대충이라도 나눠본다
        bullets = [b.strip() for b in section.split("\n") if b.strip()]

    if not bullets:
        return None, debug_context

    return " / ".join(bullets[:max_items]), debug_context


def _wait_for_list_or_entry(page, timeout_ms=15000, poll_ms=300):
    """검색 후 '목록'이 뜨는지, 검색결과가 하나뿐이라 '상세 페이지'로 바로
    넘어가는지는 미리 알 수 없다. 고정 시간을 기다리는 대신, 둘 중 무엇이
    먼저 나타나는지 짧은 간격으로 계속 확인해서 그때그때 판단한다.

    반환값: "list" | "entry" | "timeout"
    """
    elapsed = 0
    while elapsed < timeout_ms:
        if page.frame_locator("#searchIframe").locator("li").first.is_visible():
            return "list"
        if page.frame_locator("#entryIframe").locator("body").is_visible():
            return "entry"
        page.wait_for_timeout(poll_ms)
        elapsed += poll_ms
    return "timeout"


def _category_from_status_line(body_text):
    """상세 패널 글자에서 카테고리를 뽑아낸다. 두 가지 화면 형태가 섞여 있다.

    1) '...\\n브랜드명\\n카테고리\\n별점\\n4.22리뷰 3,080...' - 이 경우 '별점'이라는
       글자가 따로 있고, 그 바로 앞 줄이 카테고리다.
    2) '브랜드명\\n카테고리리뷰 3,622...' - '별점' 글자 없이 카테고리와 '리뷰'가
       바로 붙어서 나온다 (화면에는 점(·)으로 구분돼 보이지만 실제 글자에는 없다).
    """
    idx = body_text.find("별점")
    if idx != -1:
        lines = [l.strip() for l in body_text[:idx].split("\n") if l.strip()]
        if lines:
            return _strip_leading_badges(lines[-1])

    for line in body_text.split("\n"):
        if "리뷰" not in line:
            continue

        match = re.match(r"(.*?)리뷰\s*[\d,]", line)
        if not match:
            break

        category = match.group(1)
        category = re.sub(r"★[\d.]+", "", category)  # 별점이 붙어 나오면 제거
        category = category.replace("·", " ")  # 가운뎃점이 있는 경우도 대비
        category = re.sub(r"\s+", " ", category).strip()
        return _strip_leading_badges(category)

    return ""


def _read_entry_details(page, entry_frame, need_review_count, max_attempts=6):
    """목록에서 항목을 클릭해 옆에 뜬 상세 패널이 로딩되길 기다리면서,
    상태줄에서 카테고리(예: '두부요리')를 읽어온다.

    need_review_count가 True면(목록 카드 자체에 리뷰 수가 안 보였던 경우, 예:
    브랜드명만 검색했을 때) 상세 패널 글자에서 리뷰 수도 함께 읽어온다. 카테고리와
    똑같은 글자에 리뷰 수도 들어있어서 ('...별점\\n4.22리뷰 3,080...' 같은 형태),
    한 번 읽은 글자에서 같이 뽑아내면 되고 따로 더 기다릴 필요는 없다.
    """
    category, review_count = "", None
    for _ in range(max_attempts):
        try:
            body_text = entry_frame.locator("body").inner_text(timeout=1500)
        except Exception:
            body_text = ""

        if not category:
            category = _category_from_status_line(body_text)
        if need_review_count and review_count is None:
            review_count = _parse_review_count(body_text)

        if category and (not need_review_count or review_count is not None):
            return category, review_count

        page.wait_for_timeout(400)

    return category, review_count


# 'AI 브리핑' 제목은 찾았는데 그 바로 다음에 오는 게 실제 요약 내용이 아니라
# 화면 맨 아래 공통 푸터(이용약관/고객센터 등)인 경우가 있다 - 아직 요약 내용이
# 다 로딩되기 전이라 그런 것으로 보인다. 이런 경우는 "못 찾음"으로 취급해서
# 계속 기다린다 (엉뚱한 푸터 글자를 대표 리뷰로 잘못 쓰지 않기 위해).
_FOOTER_JUNK_HINTS = ("이용약관", "고객센터", "신고센터", "리뷰운영정책")


def _read_ai_briefing_snippet(page, entry_frame, max_scrolls=8):
    """목록 카드 자체에는 리뷰 문장이 안 보이는 가게들이 있다 (예약/메뉴판 위주로만
    나오는 곳 등). 이런 경우엔 상세 패널을 아래로 스크롤하면 나오는 'AI 브리핑'
    요약을 대신 가져온다.

    'AI 브리핑' 제목이 화면에 나타나는 시점과, 그 안의 실제 요약 문장이 다
    로딩되는 시점이 다를 수 있다. 그래서 1) 제목 글자가 나타날 때까지 조금씩
    스크롤해서 찾고, 2) 제목을 찾으면 그 위치로 정확히 스크롤을 맞춘 뒤,
    3) 내용이 마저 로딩되길 몇 번 더 기다려본다.
    """
    for _ in range(max_scrolls):
        try:
            body_text = entry_frame.locator("body").inner_text(timeout=1500)
        except Exception:
            body_text = ""

        if "AI 브리핑" in body_text:
            try:
                entry_frame.get_by_text("AI 브리핑").first.scroll_into_view_if_needed(timeout=2000)
            except Exception:
                pass

            for _ in range(4):
                page.wait_for_timeout(500)
                try:
                    body_text = entry_frame.locator("body").inner_text(timeout=1500)
                except Exception:
                    body_text = ""
                snippet, _ = _parse_ai_briefing(body_text)
                if snippet and not any(h in snippet for h in _FOOTER_JUNK_HINTS):
                    return snippet
            return ""

        try:
            box = entry_frame.locator("body").bounding_box()
            if box:
                page.mouse.move(box["x"] + box["width"] / 2, box["y"] + min(box["height"] / 2, 400))
        except Exception:
            pass
        page.mouse.wheel(0, 600)
        page.wait_for_timeout(400)

    return ""


def _current_list_rows(page, search_frame, min_reviews, visited, stop_event=_NO_STOP):
    """지금 화면에 있는 목록 항목들의 정보를 뽑아서 가져온다.

    (화면 밖으로 스크롤되면 항목이 DOM에서 사라지는 경우가 있어서, 스크롤하며
    보일 때마다 그때그때 뽑아둬야 한다. 나중에 한꺼번에 훑으면 이미 사라진
    항목의 정보를 놓칠 수 있다.)

    보통은(지역+키워드로 검색했을 때) 목록 카드 글자에 리뷰 수가 바로 보이므로,
    기준 리뷰 수를 넘는 항목만 클릭해서 옆에 뜨는 상세 패널로 정확한 카테고리와
    실제 지도 링크(page.url)를 가져온다 (기준 미달 항목은 클릭하지 않고 건너뛴다 -
    그래서 빠르다).

    다만 브랜드명만 검색했을 때는 네이버 지도가 목록 카드에 리뷰 수를 아예 안
    보여준다. 이 경우엔 리뷰 수를 미리 걸러낼 방법이 없으므로, 일단 클릭해서
    상세 패널에서 리뷰 수를 직접 확인한 뒤에 기준을 넘는지 판단한다.

    stop_event가 신호를 받으면(중지 버튼 등), 클릭하며 하나씩 확인하는 걸
    멈추고 지금까지 모은 것만 돌려준다.
    """
    items = search_frame.locator("li").all()
    rows = []
    for it in items:
        if stop_event.is_set():
            break
        try:
            text = it.inner_text(timeout=2000)
        except Exception:
            continue

        lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
        if not lines:
            continue
        name_guess = _strip_trailing_badges(lines[0])
        if not name_guess or name_guess in visited:
            continue

        review_count = _parse_review_count(text)
        if review_count is not None:
            # 목록 카드에 리뷰 수가 바로 보이는 일반적인 경우: 기준 미달이면 클릭 없이 건너뛴다.
            if review_count < min_reviews:
                continue
            parsed = _parse_list_item(text)
            if not parsed:
                continue
        else:
            # 목록 카드에 '리뷰' 글자 자체가 없는 경우 (예: 브랜드명만 검색) -
            # 나중에 클릭해서 상세 패널로 리뷰 수를 직접 확인해야 한다.
            if _no_review_debug_shown[0] < 3:
                print(f"  ---- (목록에 리뷰 수가 안 보여서 클릭해서 확인함 디버그) 카드 글자: {text!r}")
                _no_review_debug_shown[0] += 1
            parsed = {
                "브랜드명": name_guess,
                "리뷰수": 0,
                "대표 리뷰": _extract_review_snippet(lines),
            }

        visited.add(name_guess)

        category, map_url = "", ""
        try:
            # 카드 아무 곳이나 클릭하면 사진 영역을 눌러서 '홈'이 아니라 '사진'
            # 탭으로 들어가버리는 문제가 있었다. 파란색으로 표시되는 가게 이름
            # 링크를 직접 찾아서 그것만 클릭한다.
            try:
                it.get_by_text(parsed["브랜드명"], exact=False).first.click(timeout=3000)
            except Exception:
                it.click(position={"x": 10, "y": 10}, timeout=3000)  # 이름을 못 찾으면 예전 방식으로 대체
            page.wait_for_timeout(600)
            entry_frame = page.frame_locator("#entryIframe")
            # 혹시 그래도 '사진' 등 다른 탭으로 들어갔다면 '홈' 탭을 눌러 되돌아온다.
            try:
                home_tab = entry_frame.get_by_role("link", name="홈", exact=True)
                if home_tab.count() > 0:
                    home_tab.first.click(timeout=1500)
                    page.wait_for_timeout(400)
            except Exception:
                pass

            category, panel_review_count = _read_entry_details(
                page, entry_frame, need_review_count=(review_count is None)
            )
            map_url = page.url
            if review_count is None and panel_review_count is not None:
                parsed["리뷰수"] = panel_review_count

            # 목록 카드에서는 이름과 카테고리가 구분자 없이 붙어 나오는 경우가 있다
            # (예: '구봉만두' + '만두' -> '구봉만두만두'). 정확한 카테고리를 알고 나면,
            # 이름이 그 카테고리로 끝나는 경우 그 부분을 잘라낸다.
            if category and parsed["브랜드명"].endswith(category):
                trimmed = parsed["브랜드명"][: -len(category)].strip()
                if trimmed:
                    parsed["브랜드명"] = trimmed

            # 목록 카드에는 리뷰 문장이 아예 안 보이는 가게들이 있다(예약/메뉴 위주로만
            # 나오는 곳 등, 미쉐린 검색에서 특히 자주 보임). 이 경우 상세 패널을
            # 스크롤해서 'AI 브리핑' 요약을 대신 가져온다.
            if not parsed["대표 리뷰"]:
                ai_snippet = _read_ai_briefing_snippet(page, entry_frame)
                if ai_snippet:
                    parsed["대표 리뷰"] = ai_snippet
                if _ai_briefing_debug_shown[0] < 5:
                    print(f"  ---- (AI 브리핑 대체 디버그) {parsed['브랜드명']!r} -> {ai_snippet!r}")
                    _ai_briefing_debug_shown[0] += 1

            if _category_debug_shown[0] < 3:
                try:
                    raw = entry_frame.locator("body").inner_text(timeout=1500)
                except Exception:
                    raw = "(읽기 실패)"
                print(f"  ---- (카테고리 디버그) {parsed['브랜드명']!r} -> 카테고리={category!r}")
                print(f"  entry 패널 글자 (앞부분 500자): {raw[:500]!r}")
                _category_debug_shown[0] += 1
        except Exception:
            pass

        if parsed["리뷰수"] < min_reviews:
            continue

        parsed["카테고리"] = category
        parsed["네이버지도주소"] = map_url
        rows.append(parsed)
    return rows


def _go_to_next_page(search_frame):
    """목록 맨 아래 페이지 번호(1 2 3 4 5 >) 중 '다음 페이지' 화살표를 눌러본다.

    정확한 버튼 모양을 확인할 방법이 없어서 몇 가지 후보를 순서대로 시도한다.
    성공하면 True, 다음 페이지가 없거나 버튼을 못 찾으면 False를 돌려준다.
    """
    candidates = [
        lambda: search_frame.get_by_role("link", name=re.compile("다음")),
        lambda: search_frame.locator("a[aria-label*='다음']"),
        lambda: search_frame.locator("a:has-text('다음페이지')"),
        lambda: search_frame.locator("a.eUTV2"),  # 네이버 지도 페이지네이션에 흔히 쓰이는 클래스(추정)
    ]
    for make_locator in candidates:
        try:
            locator = make_locator()
            if locator.count() > 0 and locator.first.is_enabled():
                locator.first.click(timeout=3000)
                return True
        except Exception:
            continue
    return False


def _hover_over_list(page, search_frame):
    """가게 목록 항목 위로 마우스를 옮긴다. 목록 바깥(지도 등)에서 휠을 돌리면
    목록이 아니라 지도가 움직여버리므로, 실제 목록 카드의 화면 좌표를 찾아 그 위로
    옮긴 뒤 스크롤해야 한다. 좌표를 못 구하거나 목록 영역(화면 왼쪽) 밖으로
    나오면 대략적인 위치로 대신한다.
    """
    try:
        box = search_frame.locator("li").first.bounding_box()
        if box and box["width"] > 0 and box["x"] < 700:
            page.mouse.move(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
            return
    except Exception:
        pass
    page.mouse.move(200, 500)


def _scroll_current_page(
    page, search_frame, min_reviews, visited, stop_event=_NO_STOP, max_scrolls=25, stable_limit=4
):
    """한 페이지 안에서는 무한 스크롤로 계속 더 불러와진다. 더 이상 새 항목이
    안 늘어날 때까지(또는 max_scrolls번 스크롤할 때까지) 마우스 휠로 내리면서 모은다.

    가게 카드가 화면에 로딩되는 데 시간이 좀 걸려서, 스크롤을 조금씩 여러 번
    내리고 매번 넉넉히 기다린 뒤 다시 확인한다 (너무 빨리 "더 이상 안 늘어난다"고
    판단하면 아직 로딩 중인 걸 놓칠 수 있다).

    마우스 위치는 스크롤을 시작하기 전, 목록이 아직 그대로일 때 딱 한 번만 계산해서
    고정한다. 스크롤 도중에 다시 계산하면 화면 밖으로 스크롤된 항목의 좌표를 잘못
    짚어서 마우스가 지도 쪽으로 빠지고, 그러면 지도가 움직여버리는 문제가 있었다.
    """
    _hover_over_list(page, search_frame)

    collected = {}

    def _merge_current():
        for row in _current_list_rows(page, search_frame, min_reviews, visited, stop_event):
            collected.setdefault(row["브랜드명"], row)

    _merge_current()
    stable_rounds = 0
    for _ in range(max_scrolls):
        if stop_event.is_set():
            break
        page.mouse.wheel(0, 1500)
        page.wait_for_timeout(1200)  # 새 카드가 로딩될 시간을 넉넉히 준다
        prev_count = len(collected)
        _merge_current()

        if len(collected) <= prev_count:
            stable_rounds += 1
            if stable_rounds >= stable_limit:
                break
        else:
            stable_rounds = 0

    return list(collected.values())


def _collect_list_items_across_pages(page, max_pages, min_reviews, stop_event=_NO_STOP):
    """검색 결과 목록은 한 페이지 안에서 무한 스크롤로 꽤 많이 불러와지고,
    그 스크롤이 끝나면 '페이지 번호(1,2,3...)'로 다음 목록으로 넘어가는 구조다.
    한 페이지를 끝까지 스크롤해서 다 모은 뒤, '다음 페이지' 버튼을 눌러가며 반복한다.
    """
    search_frame = page.frame_locator("#searchIframe")
    all_rows = {}
    visited = set()  # 이미 클릭해서 상세정보를 가져온 가게 이름 (중복 클릭 방지)

    for page_num in range(1, max_pages + 1):
        if stop_event.is_set():
            break
        page.wait_for_timeout(800)  # 페이지 전환 후 목록이 그려질 시간을 준다
        page_rows = _scroll_current_page(page, search_frame, min_reviews, visited, stop_event)
        for row in page_rows:
            all_rows.setdefault(row["브랜드명"], row)
        print(f"    {page_num}페이지: 리뷰 {min_reviews}개 이상 {len(page_rows)}곳")

        if stop_event.is_set():
            break

        if page_num < max_pages:
            moved = _go_to_next_page(search_frame)
            if not moved:
                print(f"    다음 페이지 버튼을 찾지 못했습니다 ({page_num}페이지에서 멈춤).")
                break
            page.wait_for_timeout(1200)

    return list(all_rows.values())


def collect_candidates(stop_event=None):
    """config.py에 있는 지역 x 음식종류 조합마다 네이버 지도에서 검색 -> 목록 스크롤
    -> 리뷰 1000개 이상인 후보만 골라 모은다. (browser context/page는 이 함수 안에서
    새로 열고 닫는다.)

    stop_event가 신호를 받거나(launcher.py의 "중지" 버튼) Ctrl+C로 중단하면,
    처리 중이던 지점에서 멈추고 그때까지 모은 후보만 돌려준다.
    """
    stop_event = stop_event or _NO_STOP
    candidates = {}
    debug_shown = False

    # 희망지역/키워드/브랜드는 전부 선택 입력이다. 브랜드가 적혀 있으면 키워드 목록에
    # 합쳐서 똑같이 "지역 + 검색어"로 검색한다 (예: BRAND="스타벅스" -> "강남구 스타벅스").
    district = getattr(config, "DISTRICT", "").strip()
    brand = getattr(config, "BRAND", "").strip()
    search_terms = [g.strip() for g in config.PRODUCT_GROUPS if g.strip()]
    if brand:
        search_terms.append(brand)
    if not search_terms:
        # 검색어가 하나도 없으면(지역만 입력한 경우) 지역만으로 검색한다.
        search_terms = [""]
    min_reviews = config.MIN_REVIEW_COUNT or 0

    if not district and not any(search_terms):
        print("희망지역/키워드/브랜드가 모두 비어 있어서 검색할 수 없습니다.")
        return []

    with sync_playwright() as p:
        context, page = _open_browser(p)

        try:
            for group in search_terms:
                if stop_event.is_set():
                    print("중지 요청을 받아 검색을 멈춥니다. 지금까지 모은 것만 저장합니다.")
                    break

                query = " ".join(part for part in (district, group) if part)
                print(f"'{query}' 검색 중...")
                page.goto(f"https://map.naver.com/p/search/{quote(query)}", timeout=30000)

                state = _wait_for_list_or_entry(page)

                if state == "entry":
                    # 검색 결과가 통째로 1건뿐이라 목록 없이 바로 상세 페이지로 들어간 경우
                    try:
                        entry_frame = page.frame_locator("#entryIframe")
                        body_text = entry_frame.locator("body").inner_text(timeout=10000)
                        review_count = _parse_review_count(body_text)
                        name, _ = _extract_name_and_category(body_text)
                        category = _category_from_status_line(body_text)
                        if review_count is not None and review_count >= min_reviews and name:
                            candidates.setdefault(name, {
                                "브랜드명": name,
                                "키워드": group,
                                "카테고리": category,
                                "네이버지도주소": page.url,
                                "리뷰수": review_count,
                                "대표 리뷰": "",
                            })
                    except Exception as e:
                        print(f"  -> 상세 페이지를 읽지 못했습니다: {e}")
                elif state == "list":
                    rows = _collect_list_items_across_pages(
                        page,
                        max_pages=config.MAX_LIST_PAGES,
                        min_reviews=min_reviews,
                        stop_event=stop_event,
                    )

                    if not debug_shown:
                        print("  ---- (목록이 잘 읽히는지 확인용) 처음 3곳 추출 결과 ----")
                        for row in rows[:3]:
                            print("  >>>", {k: v for k, v in row.items()})
                        print("  --------------------------------------------------------")
                        debug_shown = True

                    found_in_query = 0
                    for row in rows:
                        if row["리뷰수"] < min_reviews:
                            continue
                        name = row["브랜드명"]
                        if name not in candidates:
                            candidates[name] = {
                                "브랜드명": name,
                                "키워드": group,
                                "카테고리": row["카테고리"],
                                "네이버지도주소": row["네이버지도주소"],
                                "리뷰수": row["리뷰수"],
                                "대표 리뷰": row["대표 리뷰"],
                            }
                            found_in_query += 1
                    print(f"  -> 목록 {len(rows)}개 중 리뷰 {min_reviews}개 이상 신규 {found_in_query}곳")
                else:
                    print("  -> 목록도 상세 페이지도 뜨지 않았습니다 (타임아웃). 건너뜁니다.")

                time.sleep(random.uniform(2.0, 4.0))
        except KeyboardInterrupt:
            print("\n중단 요청(Ctrl+C)을 받아 검색을 멈춥니다. 지금까지 모은 것만 저장합니다.")
        finally:
            context.close()

    return list(candidates.values())


def _try_get_extra_info(page, name):
    """가게 이름으로 다시 검색해 상세 페이지에 들어가서 주소/카테고리/지도 링크를
    읽어온다. (리뷰 수와 대표 리뷰은 목록 단계에서 이미 구했으므로 여기서는 안 읽는다.
    AI 브리핑까지 기다리는 스크롤 과정이 가게당 몇 초씩 더 걸려서 뺐다 - 필요하면
    _parse_ai_briefing()을 다시 불러 쓸 수 있다.)

    반환값: (정보 dict 또는 None, 실패 원인을 설명하는 디버그 문자열 또는 None)
    """
    # 주소 전체를 검색어로 쓰면 실제 사람이 잘 안 쓰는 특이한 검색어라 자동화로
    # 의심받기 쉬워서, 브랜드명 + 지역구 정도로 짧고 자연스러운 검색어를 사용한다.
    query = quote(f"{name} {config.DISTRICT}")
    page.goto(f"https://map.naver.com/p/search/{query}", timeout=30000)

    # 이름이 비슷한 가게가 여러 곳이면 '목록'이 뜨고, 검색결과가 하나뿐이면
    # 목록 없이 바로 '상세 페이지'로 넘어간다. 어느 쪽인지 지켜보다가
    # 목록이 뜬 경우에만 첫 번째 항목을 클릭한다.
    state = _wait_for_list_or_entry(page)
    if state == "list":
        try:
            search_frame = page.frame_locator("#searchIframe")
            search_frame.locator("li").filter(has_text=name).first.click(timeout=15000)
            page.wait_for_timeout(2500)
        except Exception:
            pass

    try:
        entry_frame = page.frame_locator("#entryIframe")
        # 상세 페이지(entryIframe)도 내용이 다 그려질 때까지 넉넉하게 기다린다.
        body_text = entry_frame.locator("body").inner_text(timeout=15000)
    except Exception as e:
        frame_names = [f.name or "(이름없음)" for f in page.frames]

        if any("captcha" in fn.lower() for fn in frame_names):
            debug = (
                "네이버가 이 접속을 '자동화 프로그램'으로 의심해서 캡차(사람 인증) 화면을 띄웠습니다.\n"
                "config.py에서 HEADLESS = False 로 바꾼 뒤 다시 실행해서, 뜨는 브라우저 창에서\n"
                "캡차를 직접 한 번 풀어주세요. 이 프로그램은 브라우저 정보를 저장해두기 때문에,\n"
                "한 번 풀고 나면 다음 실행부터는 안 떠야 정상입니다. 그래도 계속 뜨면,\n"
                "요청 속도를 더 늦추거나(가게 수를 줄이거나) 시간을 두고 다시 시도해보세요.\n"
                f"현재 페이지 URL: {page.url}"
            )
        else:
            debug = (
                f"entryIframe을 찾지 못했습니다 ({e}).\n"
                f"현재 페이지 URL: {page.url}\n"
                f"현재 페이지에 있는 프레임 이름 목록: {frame_names}"
            )
        return None, debug

    _, category = _extract_name_and_category(body_text)
    address = _district_only(_extract_address(body_text))

    info = {
        "카테고리": category,
        "주소": address,
        "네이버지도주소": page.url,
    }
    return info, None


def get_extra_info(page, name, max_attempts=2):
    """캡차가 아닌 실패는 타이밍 문제일 수 있으므로 한 번 더 재시도한다."""
    debug_info = None
    for attempt in range(1, max_attempts + 1):
        info, debug_info = _try_get_extra_info(page, name)
        if info is not None:
            return info, None

        if debug_info and "캡차" in debug_info:
            return None, debug_info

        if attempt < max_attempts:
            page.wait_for_timeout(2000)

    return None, debug_info


def _open_browser(p):
    """launch_persistent_context: 매번 새 브라우저가 아니라 browser_profile 폴더에
    쿠키/방문 기록을 저장해두고 재사용한다. 캡차를 한 번 풀면 그 기록이 남아서
    다음 실행부터는 덜 의심받는다.
    """
    context = p.chromium.launch_persistent_context(
        PROFILE_DIR,
        headless=config.HEADLESS,
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        ),
        viewport={"width": 1400, "height": 900},
        args=["--disable-blink-features=AutomationControlled"],
    )
    # 자동화 브라우저임을 알리는 대표적인 신호(navigator.webdriver)를 숨긴다.
    context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    page = context.new_page()
    return context, page


def _display_width(text):
    """열 너비/줄 수를 어림잡기 위한 글자 폭 계산. 한글 등 넓은 글자는 2칸으로 센다."""
    width = 0
    for ch in str(text):
        width += 2 if ord(ch) > 0x1100 else 1
    return width


def _resolve_output_path(output_file):
    """결과 파일을 저장할 실제 경로를 정한다.

    - 프로그램이 있는 폴더 아래 'Excel' 폴더 안에 저장한다 (폴더가 없으면 만든다).
    - 파일 이름 끝에 크롤링한 날짜(_yymmdd)를 붙인다.
    - 같은 이름의 파일이 이미 있으면(같은 조건으로 같은 날 다시 돌린 경우 등)
      '(1)', '(2)'처럼 번호를 붙여서 기존 파일을 덮어쓰지 않는다.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_dir = os.path.join(base_dir, "Excel")
    os.makedirs(excel_dir, exist_ok=True)

    name, ext = os.path.splitext(output_file)
    if not ext:
        ext = ".xlsx"
    stamped_name = f"{name}_{datetime.now().strftime('%y%m%d')}{ext}"

    candidate = os.path.join(excel_dir, stamped_name)
    counter = 1
    while os.path.exists(candidate):
        candidate = os.path.join(excel_dir, f"{name}_{datetime.now().strftime('%y%m%d')} ({counter}){ext}")
        counter += 1

    return candidate


def save_excel(rows, output_file):
    """결과를 엑셀로 저장하면서 열 너비/행 높이 자동 맞춤, 헤더 색상, 폰트를 적용한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(HEADERS)

    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    desc_col = HEADERS.index("대표 리뷰") + 1
    link_col = HEADERS.index("네이버지도주소") + 1
    center_cols = {HEADERS.index(h) + 1 for h in ("브랜드명", "키워드", "카테고리", "리뷰수")}

    # 1행(헤더): 회색 배경 + 지정 폰트 + 굵게 + 가운데 정렬
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 본문: 폰트 적용, 설명 칸은 줄바꿈 허용, 지도 링크 칸은 클릭 가능한 하이퍼링크로,
    # 브랜드명/키워드/카테고리/리뷰수 칸은 셀 안에서 가운데 정렬
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            if cell.column == link_col and cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(name=FONT_NAME, color="0563C1", underline="single")
            else:
                cell.font = Font(name=FONT_NAME)
            if cell.column == desc_col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif cell.column in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 열 너비 자동 맞춤 (글자 폭 기준 근사치 - 엑셀의 진짜 자동맞춤과 100% 같지는 않음)
    for col_index, header in enumerate(HEADERS, start=1):
        max_width = max(
            [_display_width(header)] + [_display_width(row.get(header, "")) for row in rows]
        )
        cap = 50 if header in ("대표 리뷰", "네이버지도주소") else 40
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_width + 4, cap)

    # 행 높이 자동 맞춤 (설명 칸이 몇 줄로 접힐지 어림잡아 계산)
    desc_col_width = ws.column_dimensions[get_column_letter(desc_col)].width or 50
    for row_index, row in enumerate(rows, start=2):
        desc_text = str(row.get("대표 리뷰", ""))
        lines_needed = max(1, -(-_display_width(desc_text) // int(desc_col_width)))
        ws.row_dimensions[row_index].height = max(15, lines_needed * 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)


def main(stop_event=None):
    """브랜드명/키워드/카테고리/지도링크/리뷰수/대표 리뷰을 전부 검색 목록
    단계에서만 뽑아서 바로 엑셀로 저장한다. (개별 상세 페이지는 더 이상 방문하지
    않는다 - 목록에 이미 필요한 정보가 다 있고, 하나씩 들어가면 훨씬 느려진다.
    더 정확한 정보가 필요하면 get_extra_info()를 다시 불러 쓸 수 있다.)

    stop_event(threading.Event 등)로 중간에 멈추거나 터미널에서 Ctrl+C를 눌러도,
    그때까지 모은 가게는 그대로 엑셀로 저장된다.

    반환값: (실제로 저장된 엑셀 파일의 전체 경로 또는 None, 결과 행 리스트)
    """
    print("[스크립트 버전: 2026-08-03 (브랜드명에 카테고리가 중복으로 붙는 문제 수정판)]")
    candidates = collect_candidates(stop_event=stop_event)
    if not candidates:
        print(f"리뷰 {config.MIN_REVIEW_COUNT}개 이상인 가게가 없습니다. config.py의 검색 조건을 확인하세요.")
        return None, []

    final_rows = [
        {h: c.get(h, "") for h in HEADERS}
        for c in candidates
    ]
    final_rows.sort(key=lambda r: r["리뷰수"], reverse=True)

    output_path = _resolve_output_path(config.OUTPUT_FILE)
    save_excel(final_rows, output_path)
    print(f"완료! {len(final_rows)}곳을 '{output_path}' 파일로 저장했습니다.")

    return output_path, final_rows


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n프로그램을 중단했습니다.")
