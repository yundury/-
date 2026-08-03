"""
naver_restaurant_scraper.py로 모은 결과 엑셀에서 브랜드를 체크박스로 골라,
선택한 브랜드만 구글 검색으로 추가 정보(대표번호, 본사명, 대표자명, 채용연락처)를
모아서 같은 엑셀 파일에 새 컬럼으로 덧붙이는 스크립트.

동작 순서:
1. Excel 폴더에서 가장 최근에 만들어진 결과 파일을 찾는다
   (다른 파일을 쓰고 싶으면: python enrich.py "파일 경로.xlsx")
2. 로컬 웹페이지(브라우저)를 열어서 브랜드 목록을 체크박스로 보여준다
3. "선택한 브랜드 수집 시작"을 누르면, 지점명을 뗀 브랜드 기본 이름으로 구글에서
   ① "{브랜드} 홈페이지" 검색 -> 공식 홈페이지에서 대표번호
   ② "site:saramin.co.kr {브랜드} 채용" 검색 -> 사람인 채용정보에서 본사명/대표자명/연락처
   를 찾아본다.
4. 원래 엑셀 파일에 새 컬럼으로 추가해서 저장한다.

** 중요 ** 구글 검색과 회사 홈페이지/채용사이트 글자 구조는 네이버 지도보다 훨씬
제각각이라, 이 스크립트는 완성본이 아니라 시작점입니다. 실행하면 처음 몇 곳에
대해 디버그 글자가 출력되는데, 결과가 이상하면 그 내용을 복사해서 보내주시면
정확도를 계속 다듬어갈 수 있습니다.

실행 방법: python enrich.py
"""

import glob
import html
import http.server
import json
import os
import random
import re
import socketserver
import sys
import threading
import time
import webbrowser
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

from naver_restaurant_scraper import _open_browser, FONT_NAME, HEADER_FILL, _display_width


PORT = 8765
NEW_HEADERS = ["대표번호", "본사명", "대표자명", "채용연락처"]

# 대표 번호 찾을 때 쓸 전화번호 패턴들 (02-1234-5678, 010-1234-5678, 1588-1234 등)
_PHONE_PATTERNS = [
    re.compile(r"\b0\d{1,2}-\d{3,4}-\d{4}\b"),
    re.compile(r"\b1[5-9]\d{2}-\d{4}\b"),
]

_ENRICH_DEBUG_LIMIT = 2


def _find_latest_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_dir = os.path.join(base_dir, "Excel")
    files = glob.glob(os.path.join(excel_dir, "*.xlsx"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def _load_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row_cells)))
    return rows


# ===================== 1) 체크박스 선택용 로컬 웹페이지 =====================

def _build_selection_html(rows):
    row_html = []
    for row in rows:
        name = row.get("브랜드명", "") or ""
        category = row.get("카테고리", "") or ""
        reviews = row.get("리뷰수", "") or ""
        row_html.append(
            f"<tr><td><input type='checkbox' class='brand-check' value='{html.escape(name)}'></td>"
            f"<td>{html.escape(name)}</td><td>{html.escape(str(category))}</td>"
            f"<td>{html.escape(str(reviews))}</td></tr>"
        )

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>추가 정보 수집 - 브랜드 선택</title>
<style>
  body {{ font-family: '맑은 고딕', sans-serif; background:#f4f6f8; padding: 30px; }}
  .card {{ background:#fff; border-radius: 8px; padding: 24px 32px; max-width: 800px;
           margin: 0 auto; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
  h1 {{ font-size: 20px; margin-bottom: 4px; }}
  p.desc {{ color:#666; font-size: 13px; margin-top:0; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
  th, td {{ padding: 8px 10px; border-bottom: 1px solid #eee; text-align: left; font-size: 14px; }}
  th {{ background:#f8f9fa; }}
  .toolbar {{ margin: 12px 0; }}
  button {{ background:#3b82f6; color:white; border:none; padding: 10px 20px;
            border-radius:6px; cursor:pointer; font-size:14px; }}
  button:hover {{ background:#2563eb; }}
  button.secondary {{ background:#e5e7eb; color:#333; margin-right:8px; }}
  button.secondary:hover {{ background:#d1d5db; }}
</style>
</head>
<body>
<div class="card">
  <h1>추가 정보를 수집할 브랜드를 골라주세요</h1>
  <p class="desc">체크한 브랜드만 구글 검색으로 대표번호/본사명/대표자명/채용연락처를 추가로 모읍니다.</p>
  <div class="toolbar">
    <button type="button" class="secondary" onclick="toggleAll(true)">전체 선택</button>
    <button type="button" class="secondary" onclick="toggleAll(false)">전체 해제</button>
  </div>
  <table>
    <thead><tr><th></th><th>브랜드명</th><th>카테고리</th><th>리뷰수</th></tr></thead>
    <tbody>{''.join(row_html)}</tbody>
  </table>
  <div class="toolbar">
    <button type="button" onclick="submitSelection()">선택한 브랜드 수집 시작</button>
  </div>
  <div id="status" style="margin-top:12px; color:#3b82f6; font-size:14px;"></div>
</div>
<script>
function toggleAll(checked) {{
  document.querySelectorAll('.brand-check').forEach(function(cb) {{ cb.checked = checked; }});
}}
function submitSelection() {{
  var selected = Array.from(document.querySelectorAll('.brand-check:checked')).map(function(cb) {{ return cb.value; }});
  if (selected.length === 0) {{
    alert('최소 한 곳은 선택해주세요.');
    return;
  }}
  document.getElementById('status').innerText = '요청을 보냈습니다. 이 창은 닫고 프로그램(터미널) 창에서 진행 상황을 확인하세요.';
  fetch('/submit', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{selected: selected}})
  }});
}}
</script>
</body>
</html>"""


class _Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass  # 요청마다 콘솔에 로그가 찍히지 않게 한다

    def do_GET(self):
        if self.path != "/":
            self.send_response(404)
            self.end_headers()
            return
        body = _build_selection_html(self.server.rows).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        if self.path != "/submit":
            self.send_response(404)
            self.end_headers()
            return
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length).decode("utf-8")
        data = json.loads(raw)
        self.server.selected_names = data.get("selected", [])
        self.server.submitted.set()

        body = "선택 완료!".encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def run_selection_server(rows):
    """로컬 웹서버를 띄우고 브라우저를 열어서, 사용자가 체크박스로 고를 때까지 기다린다.

    반환값: 사용자가 선택한 브랜드 이름 리스트
    """
    httpd = socketserver.TCPServer(("127.0.0.1", PORT), _Handler)
    httpd.rows = rows
    httpd.selected_names = []
    httpd.submitted = threading.Event()

    server_thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    server_thread.start()

    url = f"http://127.0.0.1:{PORT}/"
    print(f"브라우저에서 다음 주소가 자동으로 열립니다: {url}")
    print("(자동으로 안 열리면 이 주소를 직접 브라우저에 입력해주세요.)")
    webbrowser.open(url)

    httpd.submitted.wait()
    httpd.shutdown()
    return httpd.selected_names


# ===================== 2) 구글 검색으로 정보 수집 =====================

def _find_phone_number(text):
    for pattern in _PHONE_PATTERNS:
        match = pattern.search(text)
        if match:
            return match.group(0)
    return ""


def _extract_brand_base_name(name):
    """'이디야커피 강남점'처럼 지점명이 붙은 프랜차이즈면 지점명을 뗀 기본 이름을 준다.

    마지막 단어가 '점'으로 끝나면 지점명으로 보고 뗀다. 확실하지 않은 추정이다.
    """
    parts = name.split()
    if len(parts) >= 2 and parts[-1].endswith("점"):
        return " ".join(parts[:-1])
    return name


def _first_organic_result_url(page, query):
    """구글 검색 결과 중 첫 번째 일반 검색결과(구글 자체 링크 제외) 주소를 가져온다."""
    page.goto(f"https://www.google.com/search?q={quote(query)}", timeout=20000)
    page.wait_for_timeout(1500)

    try:
        links = page.locator("a[href^='http']").all()
    except Exception:
        links = []

    for link in links:
        try:
            href = link.get_attribute("href")
        except Exception:
            continue
        if not href:
            continue
        if any(domain in href for domain in ("google.com", "gstatic.com", "youtube.com")):
            continue
        return href

    return ""


def _collect_company_info(page, brand_name, debug_shown):
    """브랜드 하나에 대해 구글 검색으로 대표번호/본사명/대표자명/채용연락처를 모은다."""
    base_name = _extract_brand_base_name(brand_name)
    result = {"대표번호": "", "본사명": "", "대표자명": "", "채용연락처": ""}
    show_debug = debug_shown[0] < _ENRICH_DEBUG_LIMIT

    # 1) 공식 홈페이지에서 대표번호
    try:
        homepage_url = _first_organic_result_url(page, f"{base_name} 홈페이지")
        if homepage_url:
            page.goto(homepage_url, timeout=20000)
            page.wait_for_timeout(1200)
            body_text = page.locator("body").inner_text(timeout=5000)
            result["대표번호"] = _find_phone_number(body_text)
            if show_debug:
                print(f"    [디버그] 홈페이지 후보: {homepage_url}")
                print(f"    [디버그] 홈페이지 글자 앞부분 300자: {body_text[:300]!r}")
    except Exception as e:
        print(f"    [대표번호 수집 실패] {brand_name}: {e}")

    # 2) 사람인 채용공고에서 본사명/대표자명/연락처
    try:
        saramin_url = _first_organic_result_url(page, f"site:saramin.co.kr {base_name} 채용")
        if saramin_url:
            page.goto(saramin_url, timeout=20000)
            page.wait_for_timeout(1200)
            body_text = page.locator("body").inner_text(timeout=5000)

            company_match = re.search(r"(?:본사|회사명|기업명)\s*[:\n]\s*([^\n]+)", body_text)
            ceo_match = re.search(r"대표자?\s*[:\n]\s*([^\n]+)", body_text)

            result["본사명"] = company_match.group(1).strip() if company_match else ""
            result["대표자명"] = ceo_match.group(1).strip() if ceo_match else ""
            result["채용연락처"] = _find_phone_number(body_text)

            if show_debug:
                print(f"    [디버그] 사람인 후보: {saramin_url}")
                print(f"    [디버그] 채용정보 글자 앞부분 500자: {body_text[:500]!r}")
    except Exception as e:
        print(f"    [채용정보 수집 실패] {brand_name}: {e}")

    if show_debug:
        debug_shown[0] += 1

    return result


# ===================== 3) 엑셀에 새 컬럼으로 추가 =====================

def _update_excel_with_enrichment(path, enrichment_by_name):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    if "브랜드명" not in headers:
        print("엑셀 파일에서 '브랜드명' 컬럼을 찾지 못해 결과를 추가하지 못했습니다.")
        return
    name_col = headers.index("브랜드명") + 1

    # 이미 이전에 추가된 적이 있으면 같은 컬럼을 재사용하고, 없으면 새로 만든다
    start_col = len(headers) + 1
    new_col_index = {}
    for h in NEW_HEADERS:
        if h in headers:
            new_col_index[h] = headers.index(h) + 1
        else:
            col = start_col
            start_col += 1
            new_col_index[h] = col
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(name=FONT_NAME, bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_cells in ws.iter_rows(min_row=2):
        name = row_cells[name_col - 1].value
        info = enrichment_by_name.get(name)
        if not info:
            continue
        for h in NEW_HEADERS:
            col = new_col_index[h]
            cell = ws.cell(row=row_cells[0].row, column=col, value=info.get(h, ""))
            cell.font = Font(name=FONT_NAME)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for h in NEW_HEADERS:
        col = new_col_index[h]
        values = [ws.cell(row=r, column=col).value or "" for r in range(2, ws.max_row + 1)]
        max_width = max([_display_width(h)] + [_display_width(v) for v in values])
        ws.column_dimensions[get_column_letter(col)].width = min(max_width + 4, 30)

    wb.save(path)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else _find_latest_excel()

    if not path or not os.path.exists(path):
        print("결과 엑셀 파일을 찾지 못했습니다. 먼저 naver_restaurant_scraper.py "
              "(또는 launcher.py)를 실행해서 결과를 만들어주세요.")
        return

    print(f"'{path}' 파일에서 브랜드 목록을 불러옵니다...")
    rows = _load_rows(path)
    print(f"{len(rows)}곳을 불러왔습니다. 브라우저에서 체크박스로 골라주세요.")

    selected_names = run_selection_server(rows)
    if not selected_names:
        print("선택된 브랜드가 없어 종료합니다.")
        return
    print(f"{len(selected_names)}곳을 선택하셨습니다. 추가 정보 수집을 시작합니다...")

    enrichment_by_name = {}
    debug_shown = [0]
    with sync_playwright() as p:
        context, page = _open_browser(p)
        for i, name in enumerate(selected_names, 1):
            print(f"[{i}/{len(selected_names)}] {name} 정보 수집 중...")
            try:
                info = _collect_company_info(page, name, debug_shown)
            except Exception as e:
                print(f"  -> 실패: {e}")
                info = {"대표번호": "", "본사명": "", "대표자명": "", "채용연락처": ""}
            enrichment_by_name[name] = info
            print(f"  -> {info}")
            time.sleep(random.uniform(2.0, 4.0))
        context.close()

    _update_excel_with_enrichment(path, enrichment_by_name)
    print(f"완료! '{path}' 파일에 대표번호/본사명/대표자명/채용연락처 컬럼을 추가했습니다.")


if __name__ == "__main__":
    main()
